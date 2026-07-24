r"""Exporta el research de Panama a formatos cargables en la plataforma interna.

Uso:
    python export.py                    # exporta solo el research
    python export.py agenda-panama.json # ademas incluye el dia/hora agendado
                                        # (el backup que baja el boton "Backup")

Genera en .\export\ :
    panama-prospects.sql   upsert contra public.maximus_prospects (+ SweetWater en maximus_clients)
    panama-research.xlsx   planilla con una hoja de firmas y otra de contactos
    panama-research.csv    lo mismo en CSV (UTF-8 con BOM, abre bien en Excel)
    panama-research.json   la data estructurada, por si la consume otra app
    agenda.csv             solo si se paso el backup: quien, que dia y a que hora
"""
import csv
import json
import os
import sys

# id de la ficha local -> id real en public.maximus_prospects (plataforma interna)
IDS = {
    'fr_quantum':    'pr_c7ffcc68c3',  # Quantum Advisors
    'pr_5f0e':       'pr_b5061b0061',  # Sardona Capital S.A.
    'pr_winexco':    'pr_34ef6c3e52',  # WINEXCO SECURITIES
    'pr_singular':   'pr_1d47718b0d',  # Singular Bank
    'pr_geneva':     'pr_830d1228c8',  # Geneva Asset Management
    'pr_alpha':      'pr_4ec41113be',  # Alpha Securities
    'pr_bpsec':      'pr_287a9c7bb8',  # BP Securities
    'pr_lafise':     'pr_7550a8906b',  # Valores LAFISE
    'pr_metrobank':  'pr_7653b7dc0e',  # MetroBank
    'pr_mercantil':  'pr_5a166d9dfe',  # Mercantil Servicios de Inversion
    'pr_carlton':    'pr_c0698a3179',  # Carlton Securities
    'pr_towerbank':  'pr_6d8015e5f1',  # Towerbank International Inc.
    'pr_firmus':     'pr_c1252e466b',  # Firmus Financial
    'pr_creand':     'pr_85c6e95551',  # Creand Securities
    'pr_aliado':     'pr_e5ba89d491',  # Banco Aliado
    'pr_bgeneral':   'pr_fa2a798e43',  # Banco General
    'pr_fince':      'pr_b797a5c061',  # Fince
    'pr_seagate':    'pr_e068559756',  # SeaGate Capital
    'pr_avsec':      'pr_2ac0f1d1b5',  # AV Securities
    'pr_mmg':        'pr_d39d5fe634',  # MMG Bank
    'pr_4b38ca9672': 'pr_4b38ca9672',  # Banistmo
    'pr_054f081358': 'pr_054f081358',  # Global Bank
}
# estado actual en la plataforma (solo se usa si la fila no existiera)
ESTADO = {'contactado': 'volver_a_contactar', 'frio': 'a_contactar'}
FIT = {5: 'Muy alta', 4: 'Alta', 3: 'Media', 2: 'Baja', 1: 'Sin calificar', 0: 'Descartar'}
DIAS = {'2026-08-17': 'lunes 17/8', '2026-08-18': 'martes 18/8', '2026-08-21': 'viernes 21/8'}


def q(v):
    """Literal SQL: escapa comillas simples y mapea vacio a NULL."""
    if v is None or v == '':
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def bullets(items):
    return '\n'.join('- ' + b for b in items) if items else ''


def notas_largas(p):
    """Bloque de research para el campo notas."""
    partes = ['[Research Panama - julio 2026]']
    ficha = [x for x in (
        p.get('tipo'), p.get('grupoFin'), p.get('aum') if p.get('aum') not in ('', 'n/d') else None,
    ) if x]
    if ficha:
        partes.append(' | '.join(ficha))
    if p.get('perfil'):
        partes.append(p['perfil'])
    if p.get('fit') is not None:
        partes.append('Compatibilidad con MaximUs: ' + FIT.get(p['fit'], '') + ' (' + str(p['fit']) + '/5)')
    if p.get('pro'):
        partes.append('A favor:\n' + bullets(p['pro']))
    if p.get('contra'):
        partes.append('En contra / a validar:\n' + bullets(p['contra']))
    if p.get('alerta'):
        partes.append('ATENCION: ' + p['alerta'])
    if p.get('fuentes'):
        partes.append('Fuentes: ' + ' | '.join(f['u'] for f in p['fuentes']))
    return '\n\n'.join(partes)


def nota_plan(p):
    """Plan de accion: angulo de entrada + a quien contactar."""
    partes = []
    if p.get('angulo'):
        partes.append('Angulo de entrada: ' + p['angulo'])
    if p.get('objetivo'):
        partes.append('Objetivo de la reunion: ' + p['objetivo'])
    if p.get('contactos'):
        gente = ['A quien contactar:']
        for c in p['contactos']:
            linea = '- ' + c['n']
            if c.get('c') and c['c'] != '-':
                linea += ' (' + c['c'] + ')'
            if c.get('nota'):
                linea += ': ' + c['nota']
            gente.append(linea)
        partes.append('\n'.join(gente))
    return '\n\n'.join(partes)


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    data = json.load(open(os.path.join(base, 'data.json'), encoding='utf-8'))
    out = os.path.join(base, 'export')
    os.makedirs(out, exist_ok=True)

    # agenda opcional, desde el backup que baja la app
    agenda = {}
    if len(sys.argv) > 1:
        backup = json.load(open(sys.argv[1], encoding='utf-8'))
        for it in backup.get('items', []):
            if it.get('sched'):
                agenda[it['id']] = it['sched']

    prospectos = [p for p in data if p['grupo'] != 'cliente']
    clientes = [p for p in data if p['grupo'] == 'cliente']

    # ---------------------------------------------------------------- SQL
    sql = ["-- Research de Panama (viaje 17, 18 y 21 de agosto de 2026)",
           "-- Correr en Supabase -> SQL Editor. Actualiza contacto, notas y nota_plan.",
           "-- NO toca el estado del pipeline: eso lo decidis vos.",
           "",
           "alter table public.maximus_prospects add column if not exists nota_plan text;",
           "alter table public.maximus_clients   add column if not exists nota_plan text;",
           "",
           "insert into public.maximus_prospects (id, empresa, contacto, producto, pais, notas, nota_plan, estado) values"]
    filas = []
    for p in prospectos:
        pid = IDS.get(p['id'], p['id'])
        contacto = p.get('contacto') or ''
        if p.get('cargo') and not p['cargo'].startswith('('):
            contacto = (contacto + ' - ' + p['cargo']).strip(' -')
        filas.append('  ({}, {}, {}, {}, {}, {}, {}, {})'.format(
            q(pid), q(p['empresa']), q(contacto), q('MaximUs Pro'), q('Panama'),
            q(notas_largas(p)), q(nota_plan(p)), q(ESTADO.get(p['grupo'], 'a_contactar'))))
    sql.append(',\n'.join(filas))
    sql.append("""on conflict (id) do update set
  contacto  = excluded.contacto,
  notas     = excluded.notas,
  nota_plan = excluded.nota_plan;
""")
    sql.append("-- SweetWater Securities: alta como cliente")
    sw = next((c for c in clientes if 'Sweet' in c['empresa']), None)
    if sw:
        sql.append(
            "insert into public.maximus_clients (id, cliente, contacto, pais, servicio, accion, nota_plan)\n"
            "values ({}, {}, {}, {}, {}, {}, {})\n"
            "on conflict (id) do update set contacto = excluded.contacto, nota_plan = excluded.nota_plan;".format(
                q('cl_sweetwater'), q(sw['empresa']), q(sw.get('contacto')), q('Panama'),
                q('MaximUs'), q(sw.get('objetivo')), q(sw.get('notas'))))
    sql.append("""
-- Opcionales, descomentar si estas de acuerdo:
-- SeaGate Capital esta en liquidacion forzosa desde 2019 (Res. SMV-7-19)
-- update public.maximus_prospects set estado = 'no_les_interesa' where id = 'pr_e068559756';
-- Fince no aparece en el registro de la SMV: verificar antes de trabajarlo
-- update public.maximus_prospects set estado = 'no_les_interesa' where id = 'pr_b797a5c061';""")
    open(os.path.join(out, 'panama-prospects.sql'), 'w', encoding='utf-8').write('\n'.join(sql))

    # ---------------------------------------------------------------- tabla
    cols = [
        ('ID plataforma', lambda p: IDS.get(p['id'], p['id'])),
        ('Empresa', lambda p: p['empresa']),
        ('Categoria', lambda p: 'Con contacto previo' if p['grupo'] == 'contactado' else 'Sin contacto'),
        ('Fit', lambda p: p.get('fit')),
        ('Compatibilidad', lambda p: FIT.get(p.get('fit'), '')),
        ('Contacto', lambda p: p.get('contacto')),
        ('Cargo', lambda p: p.get('cargo')),
        ('Telefono', lambda p: p.get('tel')),
        ('Mail', lambda p: p.get('email')),
        ('Tipo', lambda p: p.get('tipo')),
        ('Licencia', lambda p: p.get('licencia')),
        ('Grupo', lambda p: p.get('grupoFin')),
        ('Activos', lambda p: p.get('aum')),
        ('Equipo', lambda p: p.get('equipo')),
        ('Foco', lambda p: p.get('foco')),
        ('Tecnologia', lambda p: p.get('tech')),
        ('Custodios', lambda p: p.get('custodios')),
        ('Que hacen', lambda p: p.get('perfil')),
        ('A favor', lambda p: bullets(p.get('pro'))),
        ('En contra', lambda p: bullets(p.get('contra'))),
        ('Angulo de entrada', lambda p: p.get('angulo')),
        ('Objetivo', lambda p: p.get('objetivo')),
        ('Prioridad', lambda p: p.get('prioridad')),
        ('Alerta', lambda p: p.get('alerta')),
        ('Direccion', lambda p: p.get('dir')),
        ('Sitio', lambda p: p.get('sitio')),
        ('Otros contactos', lambda p: '\n'.join(
            c['n'] + (' (' + c['c'] + ')' if c.get('c') and c['c'] != '-' else '') for c in p.get('contactos', []))),
        ('Fuentes', lambda p: '\n'.join(f['u'] for f in p.get('fuentes', []))),
        ('Agendado', lambda p: (lambda s: DIAS.get(s['day'], s['day']) + ' ' + s['start'] if s else '')(agenda.get(p['id']))),
    ]
    orden = sorted(prospectos, key=lambda p: (-(p.get('fit') or 0), p['empresa']))
    filas = [[fn(p) for _, fn in cols] for p in orden]

    with open(os.path.join(out, 'panama-research.csv'), 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow([c for c, _ in cols])
        w.writerows(filas)

    contactos = [[p['empresa'], c['n'], c.get('c', ''), c.get('nota', '')]
                 for p in orden for c in p.get('contactos', [])]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = 'Firmas'
        head = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', fgColor='1A2240')
        anchos = {'Que hacen': 70, 'A favor': 60, 'En contra': 60, 'Angulo de entrada': 55,
                  'Otros contactos': 40, 'Fuentes': 45, 'Empresa': 28, 'Licencia': 40, 'Grupo': 40}
        for sheet, header, rows in (
                (ws, [c for c, _ in cols], filas),
                (wb.create_sheet('Contactos'), ['Empresa', 'Nombre', 'Cargo', 'Por que'], contactos)):
            sheet.append(header)
            for r in rows:
                sheet.append(r)
            for i, name in enumerate(header, 1):
                cell = sheet.cell(row=1, column=i)
                cell.font = head
                cell.fill = fill
                sheet.column_dimensions[get_column_letter(i)].width = anchos.get(
                    name, 55 if name == 'Por que' else 22)
            sheet.freeze_panes = 'B2'
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        wb.save(os.path.join(out, 'panama-research.xlsx'))
        xlsx = 'panama-research.xlsx'
    except ImportError:
        xlsx = '(sin openpyxl, solo CSV)'

    # ---------------------------------------------------------------- JSON
    limpio = []
    for p in orden:
        d = {k: v for k, v in p.items() if k != 'sched'}
        d['id_plataforma'] = IDS.get(p['id'], p['id'])
        if agenda.get(p['id']):
            d['agendado'] = agenda[p['id']]
        limpio.append(d)
    json.dump({'viaje': 'Panama 17, 18 y 21 de agosto de 2026',
               'prospectos': limpio,
               'clientes': [{k: v for k, v in c.items() if k != 'sched'} for c in clientes]},
              open(os.path.join(out, 'panama-research.json'), 'w', encoding='utf-8'),
              ensure_ascii=False, indent=1)

    # ---------------------------------------------------------------- agenda
    if agenda:
        with open(os.path.join(out, 'agenda.csv'), 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f, delimiter=';')
            w.writerow(['Dia', 'Hora', 'Duracion (min)', 'Estado', 'Empresa', 'Contacto', 'Objetivo'])
            for p in data:
                s = agenda.get(p['id'])
                if s:
                    w.writerow([DIAS.get(s['day'], s['day']), s['start'], s['dur'], s['estado'],
                                p['empresa'], p.get('contacto', ''), p.get('objetivo', '')])

    print('Exportado a', out)
    print(' - panama-prospects.sql  ({} prospectos + SweetWater)'.format(len(prospectos)))
    print(' - ' + xlsx)
    print(' - panama-research.csv / .json')
    if agenda:
        print(' - agenda.csv ({} reuniones)'.format(len(agenda)))
    else:
        print('   (sin agenda: pasale el backup JSON como argumento para incluir dia y hora)')


if __name__ == '__main__':
    main()
